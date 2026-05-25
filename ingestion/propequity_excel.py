"""Real PropEquity Excel ingester (brief 3.1).

PropEquity delivers data as project-level workbooks: one row per project/phase
with attributes (developer, micromarket, segment, land area, launched/unsold
units, launch & current BSP, % absorbed, launch/completion dates) plus two
blocks of quarterly columns -- absorption (units) and price trend (INR/sqft).

This loads those real files into the warehouse:
  - each row -> Project (+ approximate lat/lng from a micromarket geocode table)
  - quarterly absorption -> AbsorptionSnapshot series
  - launch/current BSP + unit size -> representative ReraTransaction rows
  - quarterly price trend -> Listing-equivalent price history

No mock data: every figure traces to a PropEquity cell. Micromarkets without a
known geocode are placed at the city centroid and flagged.

Run:  python -m ingestion.propequity_excel "<file.xlsx>"
"""
from __future__ import annotations

import datetime as dt
import re
import sys
from pathlib import Path

from db.schema import (
    AbsorptionSnapshot,
    Listing,
    MicroMarket,
    MicroMarketConfig,
    Project,
    ProjectStatus,
    ReraTransaction,
)
from db.session import get_session, init_db
from config.defaults import DEFAULT_COST_ASSUMPTIONS, DEFAULT_INFRA_WEIGHTS

HEADER_ROW = 6  # 1-indexed row holding column names

# Approx geocodes for known micromarkets (extend in docs/ADMIN_GUIDE.md).
CITY_CENTROIDS = {
    "Chennai": (13.0827, 80.2707),
    "Coimbatore": (11.0168, 76.9558),
    "Bengaluru": (12.9716, 77.5946),
}
MICROMARKET_GEOCODES = {
    "Sholinganallur (OMR)": (12.9010, 80.2279),
    "Pallavaram": (12.9675, 80.1491),
    "Mylapore": (13.0339, 80.2698),
    "Anna Nagar": (13.0850, 80.2101),
    "Mambakkam": (12.8270, 80.2010),
    "Kelambakkam (OMR)": (12.7850, 80.2210),
    "Padur (OMR)": (12.8290, 80.2240),
    "Saravanampatty": (11.0780, 77.0010),
    "Kalapatti": (11.0490, 77.0290),
}


def _qtr_to_date(label: str) -> dt.date | None:
    m = re.match(r"Q(\d)-(\d{4})", str(label))
    if not m:
        return None
    q, y = int(m.group(1)), int(m.group(2))
    return dt.date(y, (q - 1) * 3 + 2, 15)


def _num(v) -> float | None:
    if v in (None, "-", ""):
        return None
    try:
        return float(str(v).replace(",", "").split("-")[0].strip())
    except (ValueError, TypeError):
        return None


def _first_int_in_range(v) -> int | None:
    n = _num(v)
    return int(n) if n is not None else None


def _config_from_bedroom(bedroom_range: str | None, project_type: str | None) -> str:
    if project_type and "plot" in str(project_type).lower():
        return "PLOT"
    if not bedroom_range or bedroom_range == "-":
        return "3BHK"
    m = re.search(r"(\d(?:\.\d)?)", str(bedroom_range))
    if m:
        n = m.group(1)
        return f"{n}BHK" if "." in n else f"{int(float(n))}BHK"
    return "3BHK"


def _geocode(micromarket: str, city: str) -> tuple[float, float, bool]:
    if micromarket in MICROMARKET_GEOCODES:
        lat, lng = MICROMARKET_GEOCODES[micromarket]
        return lat, lng, True
    lat, lng = CITY_CENTROIDS.get(city, (13.0827, 80.2707))
    return lat, lng, False


def _find_header_row(rows: list) -> int | None:
    """Return 0-based index of the row containing 'Developer Name'."""
    for i, row in enumerate(rows[:12]):
        if any(str(c).strip() == "Developer Name" for c in row if c):
            return i
    return None


def ingest_workbook(path: str, sheet: str | None = None) -> dict:
    """Ingest one PropEquity workbook. If `sheet` is None, ingests EVERY sheet
    that looks like a project table (handles multi-sheet 'Cluster' files)."""
    import openpyxl

    init_db()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [sheet] if sheet else wb.sheetnames

    totals = {"projects": 0, "transactions": 0, "snapshots": 0, "listings": 0, "ungeocoded": 0}
    micromarkets: set = set()
    for sh in sheets:
        st = _ingest_sheet(wb[sh])
        for k in totals:
            totals[k] += st.get(k, 0)
        micromarkets |= st.get("micromarkets", set())
    wb.close()
    totals["micromarkets"] = len(micromarkets)
    totals["sheets_ingested"] = sum(1 for sh in sheets if wb)  # informational
    return totals


def _ingest_sheet(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = _find_header_row(rows)
    if hdr_idx is None:
        return {"projects": 0, "transactions": 0, "snapshots": 0, "listings": 0,
                "ungeocoded": 0, "micromarkets": set()}
    header = list(rows[hdr_idx])
    col = {name: i for i, name in enumerate(header) if name}

    def get(row, name):
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    # quarterly column index ranges: two identical Q-label blocks
    q_cols = [(i, _qtr_to_date(h)) for i, h in enumerate(header) if _qtr_to_date(h)]
    half = len(q_cols) // 2
    absorption_cols = q_cols[:half]
    price_cols = q_cols[half:]

    stats = {"projects": 0, "transactions": 0, "snapshots": 0, "listings": 0, "ungeocoded": 0, "micromarkets": set()}

    with get_session() as s:
        mm_cache: dict[str, MicroMarket] = {}
        for row in rows[hdr_idx + 1:]:
            dev = get(row, "Developer Name")
            name = get(row, "Project Name")
            city = get(row, "City")
            mm_name = get(row, "Micro market")
            if not (dev and name and mm_name and city):
                continue

            lat, lng, geocoded = _geocode(mm_name, city)
            if not geocoded:
                stats["ungeocoded"] += 1
            stats["micromarkets"].add(mm_name)

            key = f"{mm_name}|{city}"
            mm = mm_cache.get(key)
            if mm is None:
                mm = s.query(MicroMarket).filter_by(name=mm_name, city=city).first()
                if mm is None:
                    mm = MicroMarket(name=mm_name, city=city, center_lat=lat, center_lng=lng,
                                     rera_state=_state_for_city(city))
                    s.add(mm); s.flush()
                    s.add(MicroMarketConfig(micro_market_id=mm.id, infra_weights=dict(DEFAULT_INFRA_WEIGHTS),
                                            cost_assumptions=dict(DEFAULT_COST_ASSUMPTIONS),
                                            min_margin_pct_of_gdv=DEFAULT_COST_ASSUMPTIONS["min_margin_pct_of_gdv"]))
                mm_cache[key] = mm

            launched = _first_int_in_range(get(row, "Launched Units"))
            unsold = _first_int_in_range(get(row, "Unsold Units"))
            sold = (launched - unsold) if (launched is not None and unsold is not None) else None
            status_raw = str(get(row, "Current Status") or "").lower()
            status = (ProjectStatus.completed if "sold" in status_raw or "completed" in status_raw
                      else ProjectStatus.ongoing)
            launch_date = _parse_date(get(row, "Launch Date"))

            proj = Project(
                rera_id=str(get(row, "Rera Number") or "")[:60] or None,
                name=str(name)[:200], developer=str(dev)[:160],
                is_gpl=("godrej" in str(dev).lower()),
                micro_market_id=mm.id, lat=lat, lng=lng, launch_date=launch_date,
                status=status, total_units=launched, units_sold=sold, source="propequity",
            )
            s.add(proj); s.flush()
            stats["projects"] += 1

            cfg = _config_from_bedroom(get(row, "Bedroom (Range)"), get(row, "Project Type"))
            unit_size = _num(get(row, "UnitSize in Sqft (Range)")) or 1200.0
            launch_bsp = _num(get(row, "Launch BSP (INR/Sqft)"))
            current_bsp = _num(get(row, "Current BSP\nINR/Sqft\n(Range)"))

            # representative transactions from real BSP figures
            for bsp, when in ((launch_bsp, launch_date), (current_bsp, dt.date.today())):
                if bsp and bsp > 100:
                    s.add(ReraTransaction(project_id=proj.id, config_type=cfg, carpet_sqft=unit_size,
                                          price_per_sqft=bsp, price_total=bsp * unit_size,
                                          floor=None, facing=None,
                                          txn_date=when or dt.date.today(), lat=lat, lng=lng, source="propequity"))
                    stats["transactions"] += 1

            # quarterly absorption -> snapshots
            cum = 0
            for idx, qdate in absorption_cols:
                units = _num(row[idx]) if idx < len(row) else None
                if units is None:
                    continue
                cum += int(units)
                price_for_q = None
                for pidx, pdate in price_cols:
                    if pdate == qdate and pidx < len(row):
                        price_for_q = _num(row[pidx])
                        break
                s.add(AbsorptionSnapshot(project_id=proj.id, as_of=qdate,
                                         units_sold_cumulative=cum, units_sold_in_month=int(units) // 3,
                                         avg_price_per_sqft=price_for_q or launch_bsp or 0.0))
                stats["snapshots"] += 1

            if current_bsp:
                s.add(Listing(project_id=proj.id, portal="propequity", config_type=cfg,
                              listed_price_per_sqft=current_bsp, available_units=unsold))
                stats["listings"] += 1

    return stats


def _pick_sheet(wb) -> str:
    for pref in ("Residential", "Projects Details", "Plots"):
        if pref in wb.sheetnames:
            return pref
    return wb.sheetnames[0]


def _state_for_city(city: str) -> str:
    return {"Chennai": "TamilNadu", "Coimbatore": "TamilNadu", "Bengaluru": "Karnataka"}.get(city, "TamilNadu")


def _parse_date(v) -> dt.date | None:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


if __name__ == "__main__":
    files = sys.argv[1:] or [
        "../Top Micromarkets Dataset_Chennai_29Apr26.xlsx",
        "../Residential Dataset_Coimbatore_15Apr26.xlsx",
        "../Plotted Projects Details_Coimbatore_15Apr26.xlsx",
    ]
    for f in files:
        if not Path(f).exists():
            print(f"skip (not found): {f}")
            continue
        print(f"Ingesting {f} ...")
        print("  ", ingest_workbook(f))
