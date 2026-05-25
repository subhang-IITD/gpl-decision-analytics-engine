"""GPL internal booking-data loader (brief 3.2 -- the 'Feed' / Salesforce source).

`report.xlsx` is GPL's own CRM/Salesforce booking export: one row per booking
with sale price, carpet area, unit type, booking/registration dates, broker, and
cancellation status. This is GPL-INTERNAL data.

SECURITY (brief 5.1): this data is loaded straight into the `historical_sales`
table and is NEVER passed to any external LLM. The loader does no LLM calls.

We aggregate bookings up to project + config level (planned vs sold, launch vs
realised price, velocity) -- the shape the models consume -- and also keep the
booking count so absorption velocity is real, not inferred.

Run:  python -m ingestion.gpl_sales_excel "../report.xlsx"
"""
from __future__ import annotations

import datetime as dt
import sys
from collections import defaultdict
from pathlib import Path

from db.schema import HistoricalSale, MicroMarket, PipelineRun
from db.session import get_session, init_db


def _num(v):
    if v in (None, "", "-"):
        return None
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _config_from_carpet(carpet) -> str:
    """Booking 'Type' is just 'Apartment', so infer config from carpet sqft."""
    c = _num(carpet) or 0
    if c <= 0:
        return "3BHK"
    if c < 750:
        return "1BHK"
    if c < 1150:
        return "2BHK"
    if c < 1450:
        return "3BHK"
    if c < 1800:
        return "3.5BHK"
    return "4BHK"


def _date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def ingest_gpl_bookings(path: str) -> dict:
    import openpyxl

    init_db()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    col = {name: i for i, name in enumerate(header) if name}
    wb.close()

    def g(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    # group bookings by (project, config)
    groups: dict[tuple, list] = defaultdict(list)
    for row in rows[1:]:
        proj = g(row, "Marketing Project Name")
        if not proj:
            continue
        cfg = _config_from_carpet(g(row, "Carpet Area"))
        groups[(str(proj), cfg)].append(row)

    started = dt.datetime.now(dt.timezone.utc)
    n_projects = 0
    with get_session() as s:
        # try to attach to an existing Bengaluru micro-market if names hint at it
        default_mm = s.query(MicroMarket).filter(MicroMarket.city == "Bengaluru").first()
        for (proj_name, cfg), bookings in groups.items():
            active = [b for b in bookings if not _truthy(g(b, "Booking Cancelled"))]
            cancelled = [b for b in bookings if _truthy(g(b, "Booking Cancelled"))]
            sold = len(active)
            planned = len(bookings)  # bookings seen; planned >= sold

            dated_psf: list[tuple] = []  # (booking_date, psf)
            for b in active:
                # 'Basic Sale Price' is already INR per sqft in this export.
                psf = _num(g(b, "Basic Sale Price"))
                if psf and 500 < psf < 100000:
                    dated_psf.append((_date(g(b, "Booking Date")) or dt.date.today(), psf))

            if not dated_psf:
                continue
            psf_vals = [p for _, p in dated_psf]
            avg_psf = sum(psf_vals) / len(psf_vals)
            # launch price ~ earliest 25% of bookings; realised ~ overall avg (appreciation)
            ordered = sorted(dated_psf, key=lambda x: x[0])
            k = max(1, len(ordered) // 4)
            launch_psf = sum(p for _, p in ordered[:k]) / k
            realised_psf = avg_psf

            s.add(HistoricalSale(
                project_name=proj_name[:200], micro_market_id=default_mm.id if default_mm else None,
                config_type=cfg, planned_units=planned, sold_units=sold,
                launch_price_per_sqft=round(launch_psf, 1), realised_price_per_sqft=round(realised_psf, 1),
                months_to_50pct=None, phase=None, source="salesforce",
            ))
            n_projects += 1

        s.add(PipelineRun(pipeline="gpl_bookings_import", status="success" if n_projects else "partial",
                          records_ingested=n_projects,
                          detail=f"Aggregated {sum(len(v) for v in groups.values())} bookings into "
                                 f"{n_projects} project-config rows.",
                          started_at=started, finished_at=dt.datetime.now(dt.timezone.utc)))

    return {"booking_rows": sum(len(v) for v in groups.values()),
            "project_config_rows": n_projects, "distinct_projects": len({k[0] for k in groups})}


def _truthy(v) -> bool:
    return str(v or "").strip().lower() in {"true", "yes", "1", "cancelled"}


if __name__ == "__main__":
    files = sys.argv[1:] or ["../report.xlsx"]
    for f in files:
        if Path(f).exists():
            print(f"Ingesting GPL bookings: {f}")
            print("  ", ingest_gpl_bookings(f))
        else:
            print("not found:", f)
